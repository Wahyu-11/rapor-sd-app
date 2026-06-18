#!/usr/bin/env python3
"""
Aplikasi Rapor SD Kelas 1-6 - Kurikulum Merdeka
Streamlit Web App untuk Input Nilai Mudah & Generate PDF Rapor Otomatis
Sesuai Panduan Pembelajaran dan Asesmen (Format Laporan Hasil Belajar SD)
"""

import streamlit as st
import pandas as pd
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import cm, mm
from reportlab.lib import colors
from reportlab.pdfgen import canvas
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import Paragraph
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_JUSTIFY, TA_RIGHT
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.lib.utils import ImageReader
from io import BytesIO
from datetime import datetime
from PIL import Image
import os
import zipfile
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# ==================== KONFIGURASI ====================
st.set_page_config(
    page_title="Aplikasi Rapor SD | Kelas 1-6",
    page_icon="📘",
    layout="wide",
    initial_sidebar_state="expanded",
    menu_items={
        'Get Help': None,
        'Report a bug': None,
        'About': "# 📘 Aplikasi Rapor SD\nVersi PWA - Bisa di-install sebagai aplikasi di Android & Desktop"
    }
)

# Custom CSS for better look
st.markdown("""
<style>
    .main-header {
        font-size: 2.2rem;
        font-weight: 700;
        color: #1a5276;
        text-align: center;
        margin-bottom: 0.5rem;
    }
    .sub-header {
        font-size: 1.1rem;
        color: #2c3e50;
        text-align: center;
        margin-bottom: 1.5rem;
    }
    .section-header {
        background-color: #eaf2f8;
        padding: 8px 12px;
        border-radius: 6px;
        font-weight: 600;
        color: #1a5276;
        margin-top: 1rem;
        margin-bottom: 0.5rem;
    }
    .stNumberInput, .stTextArea, .stTextInput, .stSelectbox {
        margin-bottom: 0.3rem;
    }
    .info-box {
        background-color: #fef9e7;
        padding: 10px;
        border-left: 4px solid #f39c12;
        border-radius: 4px;
    }
</style>
""", unsafe_allow_html=True)

# ==================== TEMA WARNA PROFESIONAL (BARU v1.4) ====================
# Tema warna menarik & profesional untuk PDF Rapor
# Memungkinkan user mengganti warna header, border, tabel agar lebih menarik sesuai branding sekolah
COLOR_THEMES = {
    "Biru Klasik (Default)": {
        "primary": "#1a5276",           # Warna utama header, judul section, border
        "table_header_bg": "#d5dbdb",   # Background header tabel nilai
        "row_alt": "#f8f9f9",           # Warna baris selang-seling tabel
        "description": "Profesional & formal, cocok untuk dokumen resmi"
    },
    "Hijau Pembelajaran": {
        "primary": "#1e8449",
        "table_header_bg": "#d5f5e3",
        "row_alt": "#eafaf1",
        "description": "Segar & edukatif, cocok untuk sekolah yang fokus lingkungan & pertumbuhan"
    },
    "Teal Modern": {
        "primary": "#117a65",
        "table_header_bg": "#d1f2eb",
        "row_alt": "#e8f6f3",
        "description": "Modern & tenang, memberikan kesan profesional yang up-to-date"
    },
    "Ungu Edukatif": {
        "primary": "#6c3483",
        "table_header_bg": "#e8daef",
        "row_alt": "#f5eef8",
        "description": "Kreatif & inspiratif, cocok untuk sekolah dengan pendekatan seni & inovasi"
    },
    "Navy Elegan": {
        "primary": "#1b4f72",
        "table_header_bg": "#d4e6f1",
        "row_alt": "#ebf5fb",
        "description": "Elegan & berwibawa, pilihan bagus untuk sekolah negeri/swasta bergengsi"
    },
}

# ==================== FUNGSI BANTU ====================
def get_fase(kelas):
    if kelas in [1, 2]:
        return "A"
    elif kelas in [3, 4]:
        return "B"
    else:
        return "C"

def get_subjects(kelas, agama_label, muatan_lokal_name="Muatan Lokal"):
    """Daftar mata pelajaran sesuai kelas (Kurikulum Merdeka).
    muatan_lokal_name: Nama kustom untuk Muatan Lokal / Kearifan Lokal Daerah (misal: 'Bahasa Jawa', 'Budaya Betawi', dll).
    """
    subjects = [
        f"Pendidikan Agama {agama_label} dan Budi Pekerti",
        "Pendidikan Pancasila",
        "Bahasa Indonesia",
        "Matematika",
    ]
    if kelas >= 3:
        subjects.append("Ilmu Pengetahuan Alam dan Sosial (IPAS)")
    subjects.extend([
        "Pendidikan Jasmani, Olahraga dan Kesehatan (PJOK)",
        "Seni Budaya",
    ])
    if kelas >= 5:
        subjects.append("Bahasa Inggris")
    subjects.append(muatan_lokal_name)
    return subjects


def get_predikat(nilai, kkm=70):
    """Menentukan predikat huruf (A/B/C/D) berdasarkan perbandingan nilai dengan KKM.
    Digunakan untuk kolom baru di tabel rapor PDF."""
    if nilai >= kkm + 15:
        return "A"
    elif nilai >= kkm:
        return "B"
    elif nilai >= kkm - 10:
        return "C"
    else:
        return "D"


def generate_deskripsi_otomatis(nama_lengkap, mapel, nilai, kkm=70):
    """Generate deskripsi capaian kompetensi yang positif, memotivasi, dan sesuai Kurikulum Merdeka.
    Deskripsi dibedakan antara nilai tinggi (di atas KKM) dan nilai rendah (di bawah KKM)."""
    if not nama_lengkap:
        nama_depan = "Ananda"
    else:
        nama_depan = nama_lengkap.split()[0].capitalize()
    
    # Level & kata-kata disesuaikan dengan KKM (nilai tinggi vs rendah)
    selisih = nilai - kkm
    
    if nilai >= kkm + 20:          # Sangat tinggi
        level = "sangat baik sekali dan membanggakan"
        capaian = "telah melampaui semua kompetensi yang diharapkan dengan sangat baik dan konsisten"
        saran = "Terus pertahankan prestasi gemilang ini! Ananda memiliki potensi untuk menjadi teladan bagi teman-teman."
    elif nilai >= kkm + 10:        # Tinggi
        level = "sangat baik"
        capaian = "telah mencapai dan melampaui kompetensi yang diharapkan dengan sangat baik"
        saran = "Pertahankan semangat belajar ini dan terus kembangkan kemampuan berpikir kritis serta kreativitas."
    elif nilai >= kkm:             # Di atas / sama dengan KKM
        level = "baik"
        capaian = "telah mencapai sebagian besar kompetensi yang diharapkan dengan baik"
        saran = "Terus berlatih dan aktif bertanya agar pemahaman semakin mendalam di semester berikutnya."
    elif nilai >= kkm - 10:        # Sedikit di bawah KKM
        level = "cukup"
        capaian = "telah menunjukkan perkembangan yang positif meskipun masih perlu sedikit bimbingan"
        saran = "Dengan latihan tambahan dan bimbingan guru/orang tua, Ananda pasti akan segera mencapai KKM."
    else:                          # Jauh di bawah KKM
        level = "masih dalam tahap berkembang"
        capaian = "sedang berusaha mencapai kompetensi dasar yang diharapkan"
        saran = "Jangan menyerah! Setiap usaha kecil akan membawa Ananda semakin dekat ke keberhasilan. Guru dan orang tua siap membantu."
    
    # Spesifik per mata pelajaran (ringkas & relevan) - tetap sama
    mapel_lower = mapel.lower()
    if "agama" in mapel_lower:
        specific = "memahami dan mengamalkan ajaran agama serta nilai-nilai budi pekerti dalam kehidupan sehari-hari"
        contoh = "Ananda sudah mampu melaksanakan ibadah dengan khusyuk dan menunjukkan sikap toleransi."
    elif "pancasila" in mapel_lower:
        specific = "memahami dan mengamalkan nilai-nilai Pancasila dalam kehidupan bermasyarakat"
        contoh = "Ananda mampu menyebutkan sila-sila Pancasila dan memberikan contoh penerapannya di sekolah dan rumah."
    elif "bahasa indonesia" in mapel_lower:
        specific = "keterampilan berbahasa Indonesia (membaca, menulis, menyimak, dan berbicara)"
        contoh = "Ananda sudah lancar membaca teks dan menulis paragraf sederhana dengan ejaan yang tepat."
    elif "matematika" in mapel_lower:
        specific = "pemahaman konsep bilangan, operasi hitung, geometri, dan pemecahan masalah"
        contoh = "Ananda mampu menyelesaikan soal cerita dan menggambar bangun datar dengan akurat."
    elif "ipas" in mapel_lower or "pengetahuan alam" in mapel_lower:
        specific = "memahami konsep IPA dan IPS serta keterampilan berpikir ilmiah dan sosial"
        contoh = "Ananda mampu mengamati fenomena alam sekitar dan menjelaskan hubungan sebab-akibat sederhana."
    elif "jasmani" in mapel_lower or "pjok" in mapel_lower:
        specific = "keterampilan gerak dasar, kebugaran jasmani, dan sportivitas"
        contoh = "Ananda aktif berpartisipasi dalam permainan dan menunjukkan sikap kerja sama yang baik."
    elif "seni" in mapel_lower:
        specific = "ekspresi kreatif melalui seni rupa, musik, tari, atau teater"
        contoh = "Ananda mampu mengekspresikan ide melalui karya seni dan menikmati berbagai bentuk seni budaya."
    elif "inggris" in mapel_lower:
        specific = "keterampilan dasar berbahasa Inggris (vocabulary, simple sentences, listening & speaking)"
        contoh = "Ananda sudah mampu memperkenalkan diri dan menyebutkan benda-benda di sekitarnya dalam bahasa Inggris sederhana."
    else:  # Muatan Lokal
        specific = "pemahaman dan pelestarian budaya serta kearifan lokal daerah"
        contoh = "Ananda menunjukkan antusiasme dalam mempelajari bahasa dan tradisi daerah."
    
    # Buat deskripsi final (lebih variatif)
    if nilai >= kkm:
        deskripsi = (
            f"{nama_depan} menunjukkan perkembangan {level} dalam {specific}. "
            f"Ananda {capaian}. {contoh} {saran} "
            f"Semangat terus, Ananda telah berhasil mencapai standar ketuntasan!"
        )
    else:
        deskripsi = (
            f"{nama_depan} menunjukkan perkembangan {level} dalam {specific}. "
            f"Ananda {capaian}. {contoh} {saran} "
            f"Dengan usaha yang konsisten, Ananda pasti akan segera tuntas dan berkembang lebih optimal!"
        )
    return deskripsi


# ==================== KETERANGAN MATA PELAJARAN (untuk bantuan input nilai) ====================
SUBJECT_KETERANGAN = {
    "Pendidikan Agama": "Membentuk karakter religius, pemahaman ajaran agama, ibadah, dan akhlak mulia sesuai keyakinan siswa.",
    "Pendidikan Pancasila": "Mengembangkan pemahaman & pengamalan nilai-nilai Pancasila, karakter bangsa, dan kehidupan bermasyarakat yang demokratis.",
    "Bahasa Indonesia": "Melatih 4 keterampilan berbahasa (menyimak, berbicara, membaca, menulis) serta menghargai sastra dan budaya Indonesia.",
    "Matematika": "Mengembangkan kemampuan berpikir logis, pemecahan masalah, konsep bilangan, geometri, pengukuran, dan statistika dasar.",
    "Ilmu Pengetahuan Alam dan Sosial (IPAS)": "Mengenal konsep IPA & IPS secara terpadu, mengembangkan keterampilan berpikir ilmiah, observasi, dan kesadaran sosial-lingkungan.",
    "Pendidikan Jasmani, Olahraga dan Kesehatan (PJOK)": "Mengembangkan kebugaran jasmani, keterampilan gerak dasar, sportivitas, kerja sama tim, dan gaya hidup sehat.",
    "Seni Budaya": "Mengekspresikan kreativitas melalui seni rupa, musik, tari, dan teater serta menghargai keberagaman seni budaya Nusantara.",
    "Bahasa Inggris": "Mengenal kosakata dasar, kalimat sederhana, listening & speaking untuk berkomunikasi dalam konteks kehidupan sehari-hari.",
    "Muatan Lokal": "Mempelajari dan melestarikan bahasa daerah, budaya, tradisi, serta kearifan lokal sesuai potensi daerah masing-masing.",
}


def create_rapor_pdf(data, pagesize=A4):
    """Generate PDF Rapor menggunakan reportlab canvas (layout profesional & rapi)
    
    pagesize: A4 (standar) atau F4 (210x330mm, lebih tinggi - cocok untuk banyak mapel tambahan)
    Perbaikan v1.3: Tabel nilai sekarang menggunakan tinggi baris DINAMIS berdasarkan jumlah baris teks aktual
    (setelah wrapping). Tidak ada lagi teks yang terpotong/hilang meskipun menambah banyak mata pelajaran ekstra
    atau deskripsi panjang. Border & garis vertikal menyesuaikan tinggi tabel yang sebenarnya.
    """
    buffer = BytesIO()
    c = canvas.Canvas(buffer, pagesize=pagesize)
    width, height = pagesize
    
    # Ambil pengaturan warna dari data (dukungan fitur baru Tema Warna & Background)
    # Fallback ke warna default jika tidak diset (backward compatible dengan versi lama)
    primary_color = data.get('primary_color', '#1a5276')
    table_header_bg = data.get('table_header_bg', '#d5dbdb')
    row_alt_color = data.get('row_alt_color', '#f8f9f9')
    page_bg_color = data.get('page_bg_color', '#FFFFFF')
    kkm = data.get('kkm', 70)
    
    # Background halaman opsional (sangat direkomendasikan warna terang seperti #F8F9FA atau #FFFFFF)
    # Berguna untuk tampilan digital yang lebih menarik; untuk cetak gunakan #FFFFFF
    if page_bg_color and str(page_bg_color).upper() != '#FFFFFF':
        c.setFillColor(colors.HexColor(page_bg_color))
        c.rect(0, 0, width, height, fill=1, stroke=0)
    
    # Margin
    left_margin = 1.5 * cm
    right_margin = 1.5 * cm
    top_margin = 1.2 * cm
    usable_width = width - left_margin - right_margin
    
    y = height - top_margin
    
    # ========== HEADER ==========
    # Kotak header (menggunakan primary_color dari tema)
    c.setFillColor(colors.HexColor(primary_color))
    c.rect(0, height - 2.8*cm, width, 2.8*cm, fill=1, stroke=0)
    
    c.setFillColor(colors.white)
    c.setFont("Helvetica-Bold", 11)
    c.drawCentredString(width/2, height - 0.9*cm, "KEMENTERIAN PENDIDIKAN, KEBUDAYAAN, RISET, DAN TEKNOLOGI")
    c.setFont("Helvetica", 9)
    c.drawCentredString(width/2, height - 1.35*cm, "REPUBLIK INDONESIA")
    
    c.setFont("Helvetica-Bold", 14)
    c.drawCentredString(width/2, height - 2.0*cm, "RAPOR PESERTA DIDIK")
    c.setFont("Helvetica-Bold", 12)
    c.drawCentredString(width/2, height - 2.45*cm, "SEKOLAH DASAR (SD)")
    
    y = height - 3.3*cm
    
    # Nama Sekolah (ikut primary_color tema)
    c.setFillColor(colors.HexColor(primary_color))
    c.setFont("Helvetica-Bold", 13)
    c.drawCentredString(width/2, y, data.get('nama_sekolah', 'SEKOLAH DASAR'))
    y -= 0.45*cm
    c.setFont("Helvetica", 9)
    c.setFillColor(colors.black)
    c.drawCentredString(width/2, y, f"NPSN: {data.get('npsn', '-')} | {data.get('alamat_sekolah', '')}")
    y -= 0.7*cm
    
    # ========== LOGO KIRI ATAS (Header) ==========
    logo_kiri_bytes = data.get('logo_kiri_bytes')
    if logo_kiri_bytes:
        try:
            logo_reader = ImageReader(BytesIO(logo_kiri_bytes))
            logo_w = 2.0 * cm
            logo_h = 2.0 * cm
            c.drawImage(logo_reader, left_margin + 0.3*cm, height - 2.55*cm,
                        width=logo_w, height=logo_h,
                        preserveAspectRatio=True, mask='auto')
        except Exception:
            pass
    
    # ========== LOGO KANAN ATAS (Header) - BARU ==========
    logo_kanan_bytes = data.get('logo_kanan_bytes')
    if logo_kanan_bytes:
        try:
            logo_reader = ImageReader(BytesIO(logo_kanan_bytes))
            logo_w = 2.0 * cm
            logo_h = 2.0 * cm
            # Posisi pojok kanan atas (simetris dengan logo kiri)
            right_logo_x = width - right_margin - logo_w - 0.3*cm
            c.drawImage(logo_reader, right_logo_x, height - 2.55*cm,
                        width=logo_w, height=logo_h,
                        preserveAspectRatio=True, mask='auto')
        except Exception:
            pass
    
    # ========== WATERMARK LOGO TRANSPARAN DI TENGAH HALAMAN ==========
    show_wm = data.get('show_watermark', False)
    logo_wm_bytes = data.get('logo_wm_bytes')
    
    # Jika tidak ada logo watermark khusus, fallback ke logo kiri (untuk backward compatibility)
    if not logo_wm_bytes and logo_kiri_bytes:
        logo_wm_bytes = logo_kiri_bytes
    
    if logo_wm_bytes and show_wm:
        try:
            logo_reader = ImageReader(BytesIO(logo_wm_bytes))
            # Ukuran watermark besar (otomatis menyesuaikan kertas A4/F4)
            wm_size = min(width * 0.58, height * 0.45)
            wm_x = (width - wm_size) / 2
            wm_y = (height - wm_size) / 2 - 1.2*cm
            
            c.saveState()
            c.setFillAlpha(0.08)
            c.drawImage(logo_reader, wm_x, wm_y,
                        width=wm_size, height=wm_size,
                        preserveAspectRatio=True, mask='auto')
            c.restoreState()
        except Exception:
            pass
    
    # Garis pemisah (ikut primary_color)
    c.setStrokeColor(colors.HexColor(primary_color))
    c.setLineWidth(1.5)
    c.line(left_margin, y, width - right_margin, y)
    y -= 0.6*cm
    
    # ========== IDENTITAS SISWA ==========
    c.setFillColor(colors.HexColor(primary_color))
    c.setFont("Helvetica-Bold", 10)
    c.drawString(left_margin, y, "IDENTITAS PESERTA DIDIK")
    y -= 0.45*cm
    
    c.setFillColor(colors.black)
    c.setFont("Helvetica", 9)
    
    # Dua kolom identitas - RAPi & PROFESIONAL: semua titik dua (colon) disejajarkan secara vertikal
    # Ini membuat tampilan lebih rapi, seperti formulir resmi (label kiri, colon sejajar, nilai setelahnya)
    col1_x = left_margin
    col2_x = left_margin + usable_width * 0.52
    
    # Data identitas (label, nilai) - pisahkan agar mudah diatur posisi colon
    identitas_kiri = [
        ("Nama Peserta Didik", data.get('nama_siswa', '-')),
        ("NISN", data.get('nisn', '-')),
        ("Kelas / Fase", f"{data.get('kelas', '-')} / Fase {data.get('fase', '-')}"),
    ]
    identitas_kanan = [
        ("Semester", data.get('semester', '-')),
        ("Tahun Ajaran", data.get('tahun_ajaran', '-')),
        ("Agama", data.get('agama', '-')),
    ]
    
    # Posisi horizontal titik dua yang KONSISTEN (supaya sejajar vertikal di setiap kolom)
    # Nilai ini sudah disesuaikan dengan lebar font Helvetica 9pt + padding rapi
    colon_x_kiri = col1_x + 3.35 * cm      # colon untuk kolom kiri (cukup untuk "Nama Peserta Didik")
    value_x_kiri = colon_x_kiri + 0.28 * cm
    
    colon_x_kanan = col2_x + 2.35 * cm     # colon untuk kolom kanan (cukup untuk "Tahun Ajaran")
    value_x_kanan = colon_x_kanan + 0.28 * cm
    
    line_height = 0.45 * cm   # sedikit lebih lega untuk keterbacaan
    
    for i, (label, val) in enumerate(identitas_kiri):
        y_pos = y - i * line_height
        c.drawString(col1_x, y_pos, label)
        c.drawString(colon_x_kiri, y_pos, ":")
        c.drawString(value_x_kiri, y_pos, str(val))
    
    for i, (label, val) in enumerate(identitas_kanan):
        y_pos = y - i * line_height
        c.drawString(col2_x, y_pos, label)
        c.drawString(colon_x_kanan, y_pos, ":")
        c.drawString(value_x_kanan, y_pos, str(val))
    
    y -= 3 * line_height + 0.15 * cm   # spacing setelah section identitas
    
    # Garis tipis
    c.setStrokeColor(colors.grey)
    c.setLineWidth(0.5)
    c.line(left_margin, y, width - right_margin, y)
    y -= 0.5*cm
    
    # ========== A. NILAI DAN CAPAIAN KOMPETENSI ==========
    c.setFillColor(colors.HexColor(primary_color))
    c.setFont("Helvetica-Bold", 10)
    c.drawString(left_margin, y, f"A. NILAI DAN CAPAIAN KOMPETENSI  |  KKM: {kkm}")
    y -= 0.5*cm
    
    # Header tabel (menggunakan table_header_bg dari tema)
    c.setFillColor(colors.HexColor(table_header_bg))
    c.rect(left_margin, y - 0.35*cm, usable_width, 0.55*cm, fill=1, stroke=0)
    
    c.setFillColor(colors.black)
    c.setFont("Helvetica-Bold", 8)
    
    # ============================================================
    # KOLOM BARU YANG RAPi & ANTI-OVERLAP (v1.6)
    # - Kolom Predikat (A/B/C/D) ditambahkan di samping Nilai
    # - Posisi absolute (cm) + gap 0.12cm dari garis vertikal
    # - Tidak ada lagi teks yang tertutup garis tabel
    # ============================================================
    col_no_x       = left_margin + 0.12*cm
    col_mapel_x    = left_margin + 0.65*cm
    mapel_w        = 5.0 * cm
    col_nilai_x    = left_margin + 5.85*cm
    nilai_w        = 1.65 * cm
    col_pred_x     = left_margin + 7.65*cm
    pred_w         = 1.65 * cm
    col_desk_x     = left_margin + 9.5*cm
    desk_w         = usable_width - (col_desk_x - left_margin) - 0.15*cm
    
    c.drawCentredString(col_no_x + 0.28*cm, y - 0.20*cm, "No")
    c.drawString(col_mapel_x, y - 0.20*cm, "Mata Pelajaran")
    c.drawCentredString(col_nilai_x + nilai_w/2, y - 0.20*cm, "Nilai")
    c.drawCentredString(col_pred_x + pred_w/2, y - 0.20*cm, "Predikat")
    c.drawString(col_desk_x, y - 0.20*cm, "Capaian Kompetensi")
    
    y -= 0.55*cm
    
    # Isi tabel nilai - DINAMIS + PREDIKAT (v1.6)
    subjects = data.get('subjects', [])
    nilai_list = data.get('nilai_list', [])
    deskripsi_list = data.get('deskripsi_list', [])
    
    row_data = []
    min_row_h = 0.90 * cm
    mapel_line_h = 0.26 * cm
    desk_line_h = 0.23 * cm
    v_padding = 0.12 * cm
    
    for mapel, nilai, desk in zip(subjects, nilai_list, deskripsi_list):
        # Wrap Mata Pelajaran
        mapel_lines = []
        words = mapel.split()
        current_line = ""
        for word in words:
            test_line = current_line + " " + word if current_line else word
            if c.stringWidth(test_line, "Helvetica", 7.5) < mapel_w:
                current_line = test_line
            else:
                if current_line:
                    mapel_lines.append(current_line)
                current_line = word
        if current_line:
            mapel_lines.append(current_line)
        
        # Wrap Deskripsi (full, dynamic)
        desk_lines = []
        words = desk.split()
        current_line = ""
        for word in words:
            test_line = current_line + " " + word if current_line else word
            if c.stringWidth(test_line, "Helvetica", 6.8) < desk_w - 0.1*cm:
                current_line = test_line
            else:
                if current_line:
                    desk_lines.append(current_line)
                current_line = word
        if current_line:
            desk_lines.append(current_line)
        
        actual_h = max(
            len(mapel_lines) * mapel_line_h + v_padding,
            len(desk_lines) * desk_line_h + v_padding,
            min_row_h
        )
        
        predikat = get_predikat(nilai, kkm)
        
        row_data.append({
            'mapel_lines': mapel_lines,
            'desk_lines': desk_lines,
            'nilai': nilai,
            'predikat': predikat,
            'height': actual_h
        })
    
    # Mulai menggambar baris data
    table_data_top = y
    for idx, rd in enumerate(row_data):
        actual_row_h = rd['height']
        
        # Background selang-seling
        if idx % 2 == 0:
            c.setFillColor(colors.HexColor(row_alt_color))
            c.rect(left_margin, y - actual_row_h + 0.08*cm, usable_width, actual_row_h, fill=1, stroke=0)
        
        c.setFillColor(colors.black)
        
        # No
        c.setFont("Helvetica", 7.5)
        c.drawCentredString(col_no_x + 0.28*cm, y - 0.30*cm, str(idx + 1))
        
        # Mata Pelajaran
        mapel_y = y - 0.22*cm
        for line in rd['mapel_lines']:
            c.drawString(col_mapel_x, mapel_y, line)
            mapel_y -= mapel_line_h
        
        # Nilai Akhir
        c.setFont("Helvetica-Bold", 9)
        c.drawCentredString(col_nilai_x + nilai_w/2, y - 0.45*cm, str(rd['nilai']))
        
        # PREDIKAT (A/B/C/D) dengan warna performa
        c.setFont("Helvetica-Bold", 9)
        if rd['predikat'] == "A":
            c.setFillColor(colors.HexColor("#1e8449"))
        elif rd['predikat'] == "B":
            c.setFillColor(colors.HexColor("#117a65"))
        elif rd['predikat'] == "C":
            c.setFillColor(colors.HexColor("#b7950b"))
        else:
            c.setFillColor(colors.HexColor("#922b21"))
        c.drawCentredString(col_pred_x + pred_w/2, y - 0.45*cm, rd['predikat'])
        c.setFillColor(colors.black)
        
        # Deskripsi
        c.setFont("Helvetica", 6.8)
        desk_y = y - 0.18*cm
        for line in rd['desk_lines']:
            c.drawString(col_desk_x, desk_y, line.strip())
            desk_y -= desk_line_h
        
        # Garis bawah hijau pada deskripsi pertama jika nilai bagus (>= kkm+5)
        # Ini mewujudkan "kata-kata bergaris pada Capaian Kompetensi sesuai nilai"
        if rd['nilai'] >= kkm + 5 and len(rd['desk_lines']) > 0:
            first_line = rd['desk_lines'][0]
            first_w = c.stringWidth(first_line, "Helvetica", 6.8)
            underline_y = (y - 0.18*cm) - 0.03*cm
            c.setStrokeColor(colors.HexColor("#27ae60"))
            c.setLineWidth(0.7)
            c.line(col_desk_x, underline_y, col_desk_x + min(first_w + 0.1*cm, desk_w - 0.15*cm), underline_y)
            c.setLineWidth(0.4)
        
        y -= actual_row_h
    
    # Border tabel utama (tegas, profesional)
    c.setStrokeColor(colors.HexColor(primary_color))
    c.setLineWidth(1.3)
    table_height = table_data_top - y + 0.55*cm
    c.rect(left_margin, y, usable_width, table_height, fill=0, stroke=1)
    
    # Garis vertikal pemisah kolom (posisi presisi + gap 0.12cm dari teks)
    c.setLineWidth(0.55)
    c.line(col_mapel_x - 0.12*cm, y, col_mapel_x - 0.12*cm, table_data_top + 0.55*cm)
    c.line(col_nilai_x - 0.12*cm, y, col_nilai_x - 0.12*cm, table_data_top + 0.55*cm)
    c.line(col_pred_x - 0.12*cm, y, col_pred_x - 0.12*cm, table_data_top + 0.55*cm)
    c.line(col_desk_x - 0.12*cm, y, col_desk_x - 0.12*cm, table_data_top + 0.55*cm)
    
    y -= 0.50*cm
    
    # ========== B/C. CAPAIAN PEMBELAJARAN (CP) & TUJUAN PEMBELAJARAN (TP) (opsional) ==========
    cp_tp = data.get('cp_tp_ringkasan', '').strip()
    cp_shown = False
    if cp_tp and cp_tp != "":
        c.setFillColor(colors.HexColor(primary_color))
        c.setFont("Helvetica-Bold", 9)
        c.drawString(left_margin, y, "B. CAPAIAN PEMBELAJARAN (CP) & TUJUAN PEMBELAJARAN (TP)")
        y -= 0.4*cm
        
        c.setFillColor(colors.black)
        c.setFont("Helvetica", 8)
        
        # Wrap teks CP/TP (lebar dinamis, sudah rapi)
        cp_lines = []
        words = cp_tp.split()
        current = ""
        for w in words:
            if c.stringWidth(current + " " + w, "Helvetica", 8) < usable_width - 0.3*cm:
                current = current + " " + w if current else w
            else:
                cp_lines.append(current)
                current = w
        if current:
            cp_lines.append(current)
        
        for line in cp_lines[:6]:  # max 6 baris
            c.drawString(left_margin, y, line)
            y -= 0.30*cm
        
        y -= 0.25*cm
        cp_shown = True
    
    # ========== C/D. KETIDAKHADIRAN ==========
    c.setFillColor(colors.HexColor(primary_color))
    c.setFont("Helvetica-Bold", 9)
    ket_label = "C. KETIDAKHADIRAN" if cp_shown else "B. KETIDAKHADIRAN"
    c.drawString(left_margin, y, ket_label)
    y -= 0.38*cm
    
    c.setFillColor(colors.black)
    c.setFont("Helvetica", 8)
    
    sakit = data.get('sakit', 0)
    izin = data.get('izin', 0)
    tanpa = data.get('tanpa_keterangan', 0)
    
    c.drawString(left_margin, y, f"Sakit              : {sakit} hari")
    c.drawString(left_margin + 5*cm, y, f"Izin               : {izin} hari")
    c.drawString(left_margin + 10*cm, y, f"Tanpa Keterangan : {tanpa} hari")
    y -= 0.55*cm
    
    # ========== D/E. CATATAN WALI KELAS ==========
    c.setFillColor(colors.HexColor(primary_color))
    c.setFont("Helvetica-Bold", 9)
    cat_label = "D. CATATAN WALI KELAS" if cp_shown else "C. CATATAN WALI KELAS"
    c.drawString(left_margin, y, cat_label)
    y -= 0.32*cm
    
    c.setFillColor(colors.black)
    c.setFont("Helvetica", 8)
    catatan = data.get('catatan_wali', 'Ananda menunjukkan perkembangan yang baik selama satu semester ini. Pertahankan semangat belajar dan tingkatkan kedisiplinan.')
    
    # Wrap catatan (lebar dinamis)
    catatan_lines = []
    words = catatan.split()
    current = ""
    for w in words:
        if c.stringWidth(current + " " + w, "Helvetica", 8) < usable_width - 0.5*cm:
            current = current + " " + w if current else w
        else:
            catatan_lines.append(current)
            current = w
    if current:
        catatan_lines.append(current)
    
    for line in catatan_lines[:5]:
        c.drawString(left_margin, y, line)
        y -= 0.30*cm
    
    y -= 0.25*cm
    
    # ========== TANDA TANGAN ==========
    c.setStrokeColor(colors.black)
    c.setLineWidth(0.5)
    
    # Posisi tanda tangan
    sig_width = usable_width / 3
    sig_y = y - 0.3*cm
    
    # Kolom 1: Wali Kelas
    c.setFont("Helvetica", 8)
    c.drawCentredString(left_margin + sig_width/2, sig_y, "Mengetahui,")
    c.drawCentredString(left_margin + sig_width/2, sig_y - 0.35*cm, "Wali Kelas")
    c.line(left_margin + 0.8*cm, sig_y - 2.2*cm, left_margin + sig_width - 0.8*cm, sig_y - 2.2*cm)
    c.setFont("Helvetica-Bold", 8)
    c.drawCentredString(left_margin + sig_width/2, sig_y - 2.5*cm, data.get('nama_wali_kelas', '........................'))
    c.setFont("Helvetica", 7)
    c.drawCentredString(left_margin + sig_width/2, sig_y - 2.75*cm, f"NIP. {data.get('nip_wali', '-')}")
    
    # Kolom 2: Kepala Sekolah
    c.setFont("Helvetica", 8)
    c.drawCentredString(left_margin + sig_width + sig_width/2, sig_y, "Kepala Sekolah")
    c.line(left_margin + sig_width + 0.8*cm, sig_y - 2.2*cm, left_margin + 2*sig_width - 0.8*cm, sig_y - 2.2*cm)
    c.setFont("Helvetica-Bold", 8)
    c.drawCentredString(left_margin + sig_width + sig_width/2, sig_y - 2.5*cm, data.get('nama_kepala', '........................'))
    c.setFont("Helvetica", 7)
    c.drawCentredString(left_margin + sig_width + sig_width/2, sig_y - 2.75*cm, f"NIP. {data.get('nip_kepala', '-')}")
    
    # Kolom 3: Orang Tua
    c.setFont("Helvetica", 8)
    c.drawCentredString(left_margin + 2*sig_width + sig_width/2, sig_y, "Orang Tua / Wali Murid")
    c.line(left_margin + 2*sig_width + 0.8*cm, sig_y - 2.2*cm, left_margin + 3*sig_width - 0.8*cm, sig_y - 2.2*cm)
    c.setFont("Helvetica-Bold", 8)
    c.drawCentredString(left_margin + 2*sig_width + sig_width/2, sig_y - 2.5*cm, "........................")
    
    # Tanggal di bawah
    y_sig = sig_y - 3.3*cm
    c.setFont("Helvetica", 8)
    tempat_tanggal = data.get('tempat_tanggal', f"{data.get('kota', '....................')}, {datetime.now().strftime('%d %B %Y')}")
    c.drawRightString(width - right_margin, y_sig, tempat_tanggal)
    
    # Footer kecil
    c.setFont("Helvetica-Oblique", 6)
    c.setFillColor(colors.grey)
    c.drawCentredString(width/2, 0.8*cm, "Dokumen ini dihasilkan secara otomatis oleh Aplikasi Rapor SD | Format sesuai Panduan Kurikulum Merdeka")
    
    c.save()
    buffer.seek(0)
    return buffer.getvalue()


# ==================== FUNGSI BATCH EXCEL (BARU) ====================
def create_excel_template():
    """Membuat template Excel siap pakai untuk input data banyak siswa"""
    wb = Workbook()
    
    # Sheet 1: Instruksi
    ws_inst = wb.active
    ws_inst.title = "INSTRUKSI"
    
    ws_inst['A1'] = "PETUNJUK PENGGUNAAN TEMPLATE RAPOR BATCH"
    ws_inst['A1'].font = Font(bold=True, size=14, color="1a5276")
    ws_inst.merge_cells('A1:G1')
    
    instructions = [
        "",
        "1. Isi data pada sheet 'Data_Siswa' (satu baris = satu siswa)",
        "2. Kolom wajib: Nama_Lengkap, NISN, Kelas (1-6), Semester, Tahun_Ajaran, Agama",
        "3. Isi nilai hanya pada kolom yang relevan dengan kelas siswa tersebut",
        "4. Kolom Nilai yang kosong akan diabaikan",
        "5. Setelah selesai, simpan file Excel ini",
        "6. Kembali ke aplikasi → Tab 'Batch Excel' → Upload file ini",
        "7. Klik 'Generate Semua PDF' → Anda akan mendapat file ZIP berisi semua rapor PDF",
        "",
        "Catatan Penting:",
        "- Agama harus ditulis persis: Islam, Kristen Protestan, Katolik, Hindu, Buddha, atau Konghucu",
        "- Semester: tulis '1 (Ganjil)' atau '2 (Genap)'",
        "- Tahun_Ajaran contoh: 2025/2026",
        "- Untuk deskripsi: sistem akan generate otomatis berdasarkan nilai (sangat disarankan)",
        "- Jika ingin deskripsi custom, gunakan mode Single Student atau edit PDF setelah dihasilkan",
        "- Nama Muatan Lokal bisa disesuaikan di UI (Single atau Batch) → misal diganti 'Bahasa Jawa' atau 'Budaya Lokal'. Kolom Excel tetap 'Nilai_Muatan_Lokal'",
        "",
        "Kolom Nilai yang tersedia:",
        "- Nilai_Agama, Nilai_Pancasila, Nilai_Bahasa_Indonesia, Nilai_Matematika",
        "- Nilai_IPAS (untuk kelas 3+), Nilai_PJOK, Nilai_Seni_Budaya",
        "- Nilai_Bahasa_Inggris (untuk kelas 5+), Nilai_Muatan_Lokal (nilai untuk mapel Muatan Lokal / kearifan lokal apapun namanya)",
    ]
    
    for i, line in enumerate(instructions, start=3):
        ws_inst[f'A{i}'] = line
        if line.startswith("Catatan") or line.startswith("Kolom Nilai"):
            ws_inst[f'A{i}'].font = Font(bold=True)
    
    ws_inst.column_dimensions['A'].width = 100
    
    # Sheet 2: Data_Siswa (Template utama)
    ws = wb.create_sheet("Data_Siswa")
    
    # Header styling
    header_fill = PatternFill(start_color="1a5276", end_color="1a5276", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=10)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Header columns
    headers = [
        "No", "Nama_Lengkap", "NISN", "Kelas", "Semester", "Tahun_Ajaran", "Agama",
        "Nilai_Agama", "Nilai_Pancasila", "Nilai_Bahasa_Indonesia", "Nilai_Matematika",
        "Nilai_IPAS", "Nilai_PJOK", "Nilai_Seni_Budaya", "Nilai_Bahasa_Inggris", "Nilai_Muatan_Lokal",
        "Sakit", "Izin", "Tanpa_Keterangan",
        "Catatan_Wali_Kelas",
        "Nama_Wali_Kelas", "NIP_Wali_Kelas",
        "Nama_Kepala_Sekolah", "NIP_Kepala_Sekolah",
        "Tempat_Tanggal",
        "Ringkasan_CP_TP"
    ]
    
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border
    
    # Sample data row (contoh)
    sample_data = [
        1, "Ahmad Fauzan", "1234567890", 3, "1 (Ganjil)", "2025/2026", "Islam",
        88, 85, 90, 82, 87, 91, 84, "", 80,   # nilai (IPAS & Inggris kosong karena contoh)
        2, 0, 0,
        "Ananda menunjukkan perkembangan yang sangat baik. Pertahankan!",
        "Siti Aminah, S.Pd.", "198505152010012003",
        "Drs. Budi Santoso, M.Pd.", "196812101990011001",
        "Jakarta, 18 Juni 2026",
        "Pada semester ini peserta didik telah mencapai Capaian Pembelajaran Fase B dengan fokus pada penguatan literasi dan numerasi."
    ]
    
    for col, value in enumerate(sample_data, start=1):
        cell = ws.cell(row=2, column=col, value=value)
        cell.border = thin_border
        cell.alignment = Alignment(wrap_text=True, vertical='center')
    
    # Set column widths
    col_widths = {
        'A': 5, 'B': 22, 'C': 14, 'D': 8, 'E': 14, 'F': 14, 'G': 18,
        'H': 12, 'I': 14, 'J': 18, 'K': 14, 'L': 12, 'M': 12, 'N': 14, 'O': 16, 'P': 14,
        'Q': 8, 'R': 8, 'S': 16,
        'T': 45,
        'U': 20, 'V': 20,
        'W': 22, 'X': 20,
        'Y': 22
    }
    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width
    
    # Row height for header
    ws.row_dimensions[1].height = 35
    
    # Add more empty rows for user to fill (20 rows ready)
    for row in range(3, 23):
        ws.cell(row=row, column=1, value=row-1)
        for col in range(1, len(headers)+1):
            ws.cell(row=row, column=col).border = thin_border
    
    # Freeze header
    ws.freeze_panes = 'A2'
    
    # Sheet 3: Contoh Hasil (optional info)
    ws3 = wb.create_sheet("Contoh_Hasil")
    ws3['A1'] = "Contoh: Setelah generate, Anda akan mendapat ZIP berisi file PDF terpisah untuk setiap siswa"
    ws3['A1'].font = Font(bold=True)
    ws3.column_dimensions['A'].width = 80
    
    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def process_batch_excel(df, nama_sekolah="SD Negeri Contoh", npsn="00000000", alamat_sekolah="", kota="Kota Contoh", 
                        logo_kiri_bytes=None, logo_kanan_bytes=None, logo_wm_bytes=None,
                        muatan_lokal_name="Muatan Lokal", show_watermark=False,
                        primary_color="#1a5276", table_header_bg="#d5dbdb", row_alt_color="#f8f9f9", page_bg_color="#FFFFFF"):
    """
    Memproses DataFrame dari Excel dan menghasilkan ZIP berisi semua PDF rapor.
    Menggunakan auto-generate deskripsi berdasarkan nilai.
    School identity diambil dari parameter UI batch (sama untuk semua siswa).
    logo_kiri_bytes, logo_kanan_bytes, logo_wm_bytes: Logo terpisah untuk kiri, kanan, dan watermark (opsional).
    muatan_lokal_name, show_watermark, primary_color, ... : Parameter lain yang sama untuk semua rapor.
    """
    zip_buffer = BytesIO()
    
    with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zipf:
        success_count = 0
        error_list = []
        
        for idx, row in df.iterrows():
            try:
                # Ambil data dasar
                nama = str(row.get('Nama_Lengkap', '')).strip()
                if not nama or nama == 'nan':
                    continue
                
                nisn = str(row.get('NISN', '')).strip()
                kelas = int(row.get('Kelas', 1))
                semester = str(row.get('Semester', '1 (Ganjil)')).strip()
                tahun_ajaran = str(row.get('Tahun_Ajaran', '2025/2026')).strip()
                agama = str(row.get('Agama', 'Islam')).strip()
                
                # Siapkan data dict seperti single mode - gunakan identitas sekolah dari parameter batch
                data = {
                    'nama_sekolah': nama_sekolah,
                    'npsn': npsn,
                    'alamat_sekolah': alamat_sekolah,
                    'kota': kota,
                    'nama_siswa': nama,
                    'nisn': nisn,
                    'kelas': kelas,
                    'fase': get_fase(kelas),
                    'semester': semester,
                    'tahun_ajaran': tahun_ajaran,
                    'agama': agama,
                    # Logo terpisah v1.5
                    'logo_kiri_bytes': logo_kiri_bytes,
                    'logo_kanan_bytes': logo_kanan_bytes,
                    'logo_wm_bytes': logo_wm_bytes,
                    'show_watermark': show_watermark,
                    # Warna tema batch
                    'primary_color': primary_color,
                    'table_header_bg': table_header_bg,
                    'row_alt_color': row_alt_color,
                    'page_bg_color': page_bg_color,
                }
                
                # Bangun daftar mapel & nilai berdasarkan kelas
                subjects = get_subjects(kelas, agama, muatan_lokal_name)
                nilai_list = []
                deskripsi_list = []
                
                # Mapping nama kolom Excel ke mapel (kolom Excel tetap "Nilai_Muatan_Lokal" untuk kesederhanaan)
                mapel_to_col = {
                    f"Pendidikan Agama {agama} dan Budi Pekerti": "Nilai_Agama",
                    "Pendidikan Pancasila": "Nilai_Pancasila",
                    "Bahasa Indonesia": "Nilai_Bahasa_Indonesia",
                    "Matematika": "Nilai_Matematika",
                    "Ilmu Pengetahuan Alam dan Sosial (IPAS)": "Nilai_IPAS",
                    "Pendidikan Jasmani, Olahraga dan Kesehatan (PJOK)": "Nilai_PJOK",
                    "Seni Budaya": "Nilai_Seni_Budaya",
                    "Bahasa Inggris": "Nilai_Bahasa_Inggris",
                    muatan_lokal_name: "Nilai_Muatan_Lokal",
                }
                
                for mapel in subjects:
                    col_name = mapel_to_col.get(mapel, "")
                    nilai = row.get(col_name, 0)
                    if pd.isna(nilai) or nilai == "" or nilai == 0:
                        nilai = 75  # default jika kosong
                    
                    nilai = int(float(nilai))
                    nilai_list.append(nilai)
                    
                    # Generate deskripsi otomatis
                    desk = generate_deskripsi_otomatis(nama, mapel, nilai)
                    deskripsi_list.append(desk)
                
                data['subjects'] = subjects
                data['nilai_list'] = nilai_list
                data['deskripsi_list'] = deskripsi_list
                
                # Kehadiran
                data['sakit'] = int(row.get('Sakit', 0) or 0)
                data['izin'] = int(row.get('Izin', 0) or 0)
                data['tanpa_keterangan'] = int(row.get('Tanpa_Keterangan', 0) or 0)
                
                # Catatan & Tanda tangan
                catatan = str(row.get('Catatan_Wali_Kelas', '')).strip()
                if not catatan or catatan == 'nan':
                    catatan = "Ananda menunjukkan perkembangan yang baik selama semester ini. Pertahankan semangat belajar dan tingkatkan kedisiplinan."
                data['catatan_wali'] = catatan
                
                data['nama_wali_kelas'] = str(row.get('Nama_Wali_Kelas', '........................')).strip()
                data['nip_wali'] = str(row.get('NIP_Wali_Kelas', '-')).strip()
                data['nama_kepala'] = str(row.get('Nama_Kepala_Sekolah', '........................')).strip()
                data['nip_kepala'] = str(row.get('NIP_Kepala_Sekolah', '-')).strip()
                
                tempat_tanggal = str(row.get('Tempat_Tanggal', '')).strip()
                if not tempat_tanggal or tempat_tanggal == 'nan':
                    tempat_tanggal = f"{data['kota']}, {datetime.now().strftime('%d %B %Y')}"
                data['tempat_tanggal'] = tempat_tanggal
                
                # CP & TP ringkasan (opsional dari Excel)
                cp_tp = str(row.get('Ringkasan_CP_TP', '')).strip()
                if not cp_tp or cp_tp == 'nan':
                    cp_tp = f"Pada semester ini peserta didik telah mencapai Capaian Pembelajaran Fase {get_fase(kelas)} pada mata pelajaran yang diajarkan."
                data['cp_tp_ringkasan'] = cp_tp
                
                # Generate PDF
                pdf_bytes = create_rapor_pdf(data)
                
                # Nama file PDF
                safe_nama = nama.replace(" ", "_").replace("/", "-")
                pdf_filename = f"Rapor_{safe_nama}_Kelas{kelas}_{semester.replace(' ', '')}_{tahun_ajaran.replace('/', '-')}.pdf"
                
                zipf.writestr(pdf_filename, pdf_bytes)
                success_count += 1
                
            except Exception as e:
                error_list.append(f"Baris {idx+2}: {nama} - Error: {str(e)}")
        
        # Tambahkan file ringkasan
        summary = f"""RINGKASAN GENERATE RAPOR BATCH
================================
Tanggal Generate : {datetime.now().strftime('%d %B %Y %H:%M')}
Total Siswa Berhasil : {success_count}
Total Error : {len(error_list)}

"""
        if error_list:
            summary += "DAFTAR ERROR:\n" + "\n".join(error_list)
        
        zipf.writestr("RINGKASAN_GENERATE.txt", summary.encode('utf-8'))
    
    zip_buffer.seek(0)
    return zip_buffer.getvalue(), success_count, error_list


# ==================== UI UTAMA ====================
st.markdown('<h1 class="main-header">📘 APLIKASI RAPOR SD</h1>', unsafe_allow_html=True)
st.markdown('<p class="sub-header">Kelas 1–6 | Kurikulum Merdeka | Input Mudah • Deskripsi Otomatis • PDF Siap Unduh</p>', unsafe_allow_html=True)

# Sidebar
with st.sidebar:
    st.header("⚙️ Pengaturan")
    st.info("Aplikasi ini mendukung semua mata pelajaran SD sesuai Kurikulum Merdeka (termasuk IPAS mulai kelas 3 & Bahasa Inggris kelas 5-6).")
    
    st.markdown("---")
    st.subheader("Petunjuk Singkat")
    st.markdown("""
    1. Isi **Identitas Sekolah** & **Siswa**
    2. Pilih **Kelas** & **Agama** (untuk mata pelajaran Agama)
    3. Input **Nilai Akhir** (0-100)
    4. Edit atau gunakan **Deskripsi Otomatis** yang sudah sesuai kaidah Kurikulum Merdeka
    5. Lengkapi kehadiran & catatan
    6. Klik **Generate PDF** → Langsung unduh!
    """)
    
    st.markdown("---")
    st.caption("Versi 1.4 | 🎨 Tema Warna & Background | Format sesuai Panduan Kurikulum Merdeka")

    # === PWA INSTALLATION GUIDE ===
    with st.expander("📱 Install sebagai Aplikasi di Android (PWA)", expanded=False):
        st.markdown("""
        **Cara Install di HP Android:**
        
        1. Deploy aplikasi ini ke internet (bisa pakai Streamlit Cloud gratis)
        2. Buka link aplikasi menggunakan **Chrome** di HP Android
        3. Ketuk ikon **⋮** (tiga titik) di pojok kanan atas
        4. Pilih **"Add to Home screen"** atau **"Install app"**
        5. Konfirmasi → Ikon aplikasi akan muncul di layar utama HP Anda
        
        Setelah di-install, aplikasi akan terbuka dalam mode fullscreen seperti aplikasi native.
        """)

# Tab utama
tab1, tab2, tab3 = st.tabs(["📝 Input Data Rapor (Single)", "📦 Batch Banyak Siswa (Excel)", "📖 Panduan & Contoh"])

with tab1:
    # === IDENTITAS SEKOLAH ===
    st.markdown('<div class="section-header">🏫 IDENTITAS SEKOLAH</div>', unsafe_allow_html=True)
    
    col1, col2 = st.columns(2)
    with col1:
        nama_sekolah = st.text_input("Nama Sekolah Dasar", value="SD Negeri 01 Contoh", key="nama_sekolah")
        npsn = st.text_input("NPSN", value="12345678", key="npsn")
    with col2:
        alamat_sekolah = st.text_input("Alamat Sekolah", value="Jl. Pendidikan No. 1, Kota Contoh", key="alamat")
        kota = st.text_input("Kota/Kabupaten", value="Kota Contoh", key="kota")
    
    # ==================== LOGO UPLOAD (3 TEMPAT) - BARU v1.5 ====================
    st.markdown("**🖼️ Upload Logo (Maksimal 3 Logo)**")
    st.caption("Anda bisa mengunggah logo berbeda untuk posisi yang berbeda. Logo sebaiknya berlatar belakang transparan atau putih bersih.")
    
    col_l1, col_l2, col_l3 = st.columns(3)
    
    with col_l1:
        logo_kiri = st.file_uploader(
            "📍 Logo Kiri Atas (Header)", 
            type=["png", "jpg", "jpeg"],
            key="logo_kiri",
            help="Logo kecil yang muncul di pojok kiri atas header rapor (ukuran ~2x2 cm)."
        )
    
    with col_l2:
        logo_kanan = st.file_uploader(
            "📍 Logo Kanan Atas (Header)", 
            type=["png", "jpg", "jpeg"],
            key="logo_kanan",
            help="Logo kecil yang muncul di pojok kanan atas header rapor (misalnya logo Kemendikbud atau emblem lain)."
        )
    
    with col_l3:
        logo_wm = st.file_uploader(
            "💧 Logo Watermark Tengah", 
            type=["png", "jpg", "jpeg"],
            key="logo_wm",
            help="Logo besar yang akan ditampilkan samar di tengah halaman sebagai watermark. Ukuran otomatis menyesuaikan."
        )
    
    show_watermark = st.checkbox(
        "💧 Tampilkan Logo Watermark Tengah di TENGAH halaman (samar/transparan)",
        value=True if logo_wm else False,
        key="show_watermark",
        disabled=logo_wm is None,
        help="Jika dicentang, logo dari kolom 'Logo Watermark Tengah' akan muncul besar dan samar di tengah halaman sebagai background watermark."
    )
    
    # ==================== FITUR BARU: TEMA WARNA & BACKGROUND PDF ====================
    with st.expander("🎨 Kustomisasi Warna & Tema Rapor (Baru - v1.4)", expanded=False):
        st.markdown("""
        **Pilih tema warna untuk membuat rapor lebih menarik & sesuai branding sekolah Anda.**
        - Tema yang disediakan sudah dirancang profesional & mudah dibaca.
        - Header, border, judul section, dan styling tabel akan menyesuaikan warna pilihan Anda.
        - Untuk hasil terbaik saat **cetak**, pilih warna yang tidak terlalu gelap.
        - Fitur **Background Halaman** hanya disarankan untuk tampilan digital (warna sangat terang).
        """)
        
        theme_options = list(COLOR_THEMES.keys()) + ["Custom (Pilih Warna Manual)"]
        theme_choice = st.selectbox(
            "Pilih Tema Warna Profesional",
            options=theme_options,
            index=0,
            key="theme_choice",
            help="Tema akan diterapkan langsung ke PDF yang dihasilkan."
        )
        
        if theme_choice == "Custom (Pilih Warna Manual)":
            st.caption("Atur warna sesuai keinginan Anda. Disarankan primary/header berwarna gelap agar teks putih kontras.")
            col_c1, col_c2 = st.columns(2)
            with col_c1:
                primary_color = st.color_picker("Warna Utama Header, Border & Judul", "#1a5276", key="custom_primary")
                table_header_bg = st.color_picker("Warna Background Header Tabel", "#d5dbdb", key="custom_table_header")
            with col_c2:
                row_alt_color = st.color_picker("Warna Baris Selang-seling Tabel", "#f8f9f9", key="custom_row_alt")
                page_bg_color = st.color_picker("Warna Background Halaman (opsional, sangat terang)", "#FFFFFF", key="custom_page_bg")
                st.caption("Contoh background menarik: #F4F6F7 (abu terang), #EBF5FB (biru muda), #FEF9E7 (krem)")
        else:
            theme = COLOR_THEMES[theme_choice]
            primary_color = theme["primary"]
            table_header_bg = theme["table_header_bg"]
            row_alt_color = theme["row_alt"]
            page_bg_color = "#FFFFFF"  # default putih untuk tema preset (bisa diubah di custom)
            st.info(f"**{theme_choice}**: {theme['description']}")
            # Tampilkan preview warna
            st.markdown(f"""
            <div style="display:flex; gap:10px; margin-top:8px;">
                <div style="background:{primary_color}; width:60px; height:25px; border-radius:4px; border:1px solid #ccc;"></div>
                <div style="background:{table_header_bg}; width:60px; height:25px; border-radius:4px; border:1px solid #ccc;"></div>
                <div style="background:{row_alt_color}; width:60px; height:25px; border-radius:4px; border:1px solid #ccc;"></div>
            </div>
            <small>Primary • Table Header • Row Alt</small>
            """, unsafe_allow_html=True)
    
    # Simpan pilihan warna ke session_state agar bisa diakses saat generate
    st.session_state['primary_color'] = primary_color
    st.session_state['table_header_bg'] = table_header_bg
    st.session_state['row_alt_color'] = row_alt_color
    st.session_state['page_bg_color'] = page_bg_color
    
    # === IDENTITAS SISWA ===
    st.markdown('<div class="section-header">👤 IDENTITAS PESERTA DIDIK</div>', unsafe_allow_html=True)
    
    col1, col2, col3 = st.columns([2, 1, 1])
    with col1:
        nama_siswa = st.text_input("Nama Lengkap Siswa", value="Ahmad Fauzan", key="nama_siswa")
    with col2:
        nisn = st.text_input("NISN", value="1234567890", key="nisn")
    with col3:
        kelas = st.selectbox("Kelas", options=list(range(1, 7)), index=2, key="kelas")  # default kelas 3
    
    col1, col2, col3 = st.columns(3)
    with col1:
        semester = st.selectbox("Semester", ["1 (Ganjil)", "2 (Genap)"], index=0, key="semester")
    with col2:
        tahun_ajaran = st.text_input("Tahun Ajaran", value="2025/2026", key="tahun_ajaran")
    with col3:
        agama = st.selectbox("Agama (untuk mapel Agama)", 
                             ["Islam", "Kristen Protestan", "Katolik", "Hindu", "Buddha", "Konghucu"], 
                             index=0, key="agama")
    
    fase = get_fase(kelas)
    st.caption(f"Fase otomatis: **Fase {fase}** (sesuai kelas {kelas})")
    
    # Pilihan Nama Muatan Lokal (BARU - fleksibel sesuai daerah/sekolah)
    muatan_lokal_nama = st.text_input(
        "📍 Nama Muatan Lokal / Kearifan Lokal Daerah",
        value="Muatan Lokal",
        key="muatan_lokal_nama",
        help="Sesuaikan dengan mata pelajaran Muatan Lokal di sekolah Anda (contoh: Bahasa Jawa, Bahasa Sunda, Budaya Betawi, Seni Tradisi Lokal, Kearifan Lokal, dll). Label ini akan digunakan di tabel nilai rapor. Biarkan default jika sekolah menggunakan istilah 'Muatan Lokal'."
    )
    
    # === MATA PELAJARAN & NILAI ===
    st.markdown('<div class="section-header">📊 NILAI AKHIR & CAPAIAN KOMPETENSI</div>', unsafe_allow_html=True)
    
    # Input KKM (global untuk rapor ini)
    col_kkm1, col_kkm2 = st.columns([0.4, 0.6])
    with col_kkm1:
        kkm = st.number_input(
            "KKM (Kriteria Ketuntasan Minimal)", 
            min_value=50, max_value=100, value=70, step=5,
            key="kkm_input",
            help="Nilai minimal yang harus dicapai siswa untuk dinyatakan tuntas pada mata pelajaran."
        )
    with col_kkm2:
        st.caption("Predikat dan status ketuntasan akan disesuaikan berdasarkan KKM ini. Contoh: Jika KKM=70, nilai ≥85 biasanya mendapat predikat B atau A.")
    
    agama_label = agama
    subjects = get_subjects(kelas, agama_label, muatan_lokal_nama)
    
    st.caption(f"📌 {len(subjects)} mata pelajaran untuk Kelas {kelas} (Fase {fase}) — Muatan Lokal ditampilkan sebagai: **{muatan_lokal_nama}**")
    
    # Inisialisasi session state untuk nilai & deskripsi
    if 'nilai_data' not in st.session_state or st.session_state.get('last_kelas') != kelas or st.session_state.get('last_agama') != agama or st.session_state.get('last_muatan_lokal') != muatan_lokal_nama:
        st.session_state.nilai_data = {}
        st.session_state.last_kelas = kelas
        st.session_state.last_agama = agama
        st.session_state.last_muatan_lokal = muatan_lokal_nama
    
    # Inisialisasi mata pelajaran tambahan (bisa ditambah/hapus oleh user)
    if 'extra_subjects' not in st.session_state:
        st.session_state.extra_subjects = []
    if 'extra_nilai' not in st.session_state:
        st.session_state.extra_nilai = {}
    if 'extra_deskripsi' not in st.session_state:
        st.session_state.extra_deskripsi = {}
    
    nilai_list = []
    deskripsi_list = []
    
    # Input per mata pelajaran
    for i, mapel in enumerate(subjects):
        with st.container(border=True):
            col_nilai, col_desk = st.columns([0.22, 0.78])
            
            with col_nilai:
                default_nilai = st.session_state.nilai_data.get(mapel, 82)
                nilai = st.number_input(
                    f"Nilai Akhir", 
                    min_value=0, max_value=100, value=default_nilai, 
                    key=f"nilai_{i}",
                    help="Masukkan nilai 0-100"
                )
                nilai_list.append(nilai)
                st.session_state.nilai_data[mapel] = nilai
                
                # Predikat otomatis berdasarkan KKM
                if nilai >= kkm + 15:
                    predikat = "A (Sangat Baik)"
                elif nilai >= kkm:
                    predikat = "B (Baik)"
                elif nilai >= kkm - 10:
                    predikat = "C (Cukup)"
                else:
                    predikat = "D (Perlu Bimbingan)"
                st.caption(f"Predikat: **{predikat}** | KKM: {kkm}")
            
            with col_desk:
                # Generate atau edit deskripsi
                default_desk = generate_deskripsi_otomatis(nama_siswa, mapel, nilai, kkm)
                if mapel not in st.session_state.nilai_data:
                    st.session_state.nilai_data[mapel] = nilai
                
                deskripsi = st.text_area(
                    f"Deskripsi Capaian Kompetensi", 
                    value=default_desk, 
                    height=75, 
                    key=f"desk_{i}",
                    help="Edit sesuai observasi nyata siswa. Bahasa sudah positif & memotivasi."
                )
                deskripsi_list.append(deskripsi)
            
            # Keterangan singkat mata pelajaran (di bawah kolom)
            mapel_key = next((k for k in SUBJECT_KETERANGAN.keys() if k in mapel), None)
            if mapel_key:
                st.caption(f"ℹ️ **{mapel_key}**: {SUBJECT_KETERANGAN[mapel_key]}")
    
    # Tombol regenerate semua deskripsi
    if st.button("🔄 Regenerate Semua Deskripsi Otomatis", type="secondary"):
        for i, mapel in enumerate(subjects):
            new_desk = generate_deskripsi_otomatis(nama_siswa, mapel, nilai_list[i], kkm)
            st.session_state[f"desk_{i}"] = new_desk
        st.rerun()
    
    st.info("💡 Deskripsi sudah disesuaikan dengan level nilai dan mata pelajaran. Anda bebas mengedit untuk lebih akurat.")
    
    # ========== FITUR TAMBAH / HAPUS MATA PELAJARAN (BARU) ==========
    with st.expander("➕ Tambah / Hapus Mata Pelajaran Tambahan (Opsional)", expanded=False):
        st.caption("Gunakan fitur ini jika ada mata pelajaran tambahan di luar standar Kurikulum Merdeka untuk kelas tersebut (misalnya Muatan Lokal khusus atau kegiatan unggulan sekolah).")
        
        col_add1, col_add2 = st.columns([0.7, 0.3])
        with col_add1:
            new_mapel_name = st.text_input("Nama Mata Pelajaran Baru", placeholder="Contoh: Bahasa Daerah / Keterampilan", key="new_mapel_input")
        with col_add2:
            if st.button("➕ Tambah Mapel", type="secondary", use_container_width=True):
                if new_mapel_name.strip() and new_mapel_name.strip() not in st.session_state.extra_subjects:
                    st.session_state.extra_subjects.append(new_mapel_name.strip())
                    st.session_state.extra_nilai[new_mapel_name.strip()] = 75
                    st.session_state.extra_deskripsi[new_mapel_name.strip()] = generate_deskripsi_otomatis(nama_siswa, new_mapel_name.strip(), 75, kkm)
                    st.rerun()
                elif new_mapel_name.strip() in st.session_state.extra_subjects:
                    st.warning("Mata pelajaran sudah ada!")
        
        if st.session_state.extra_subjects:
            st.markdown("**Mata Pelajaran Tambahan yang Sudah Ditambahkan:**")
            for idx, extra_mapel in enumerate(st.session_state.extra_subjects):
                with st.container(border=True):
                    col_e1, col_e2, col_e3 = st.columns([0.55, 0.25, 0.2])
                    with col_e1:
                        extra_nilai_val = st.number_input(
                            f"Nilai {extra_mapel}", 
                            min_value=0, max_value=100, 
                            value=st.session_state.extra_nilai.get(extra_mapel, 75),
                            key=f"extra_nilai_{idx}"
                        )
                        st.session_state.extra_nilai[extra_mapel] = extra_nilai_val
                        
                        # Predikat otomatis untuk extra (berdasarkan KKM)
                        if extra_nilai_val >= kkm + 15:
                            extra_pred = "A"
                        elif extra_nilai_val >= kkm:
                            extra_pred = "B"
                        elif extra_nilai_val >= kkm - 10:
                            extra_pred = "C"
                        else:
                            extra_pred = "D"
                        st.caption(f"Predikat: **{extra_pred}** | KKM: {kkm}")
                    
                    with col_e2:
                        extra_desk = st.text_area(
                            f"Deskripsi {extra_mapel}", 
                            value=st.session_state.extra_deskripsi.get(extra_mapel, ""),
                            height=60,
                            key=f"extra_desk_{idx}"
                        )
                        st.session_state.extra_deskripsi[extra_mapel] = extra_desk
                    
                    with col_e3:
                        if st.button("🗑️ Hapus", key=f"del_extra_{idx}", type="secondary"):
                            # Hapus dari semua storage
                            if extra_mapel in st.session_state.extra_subjects:
                                st.session_state.extra_subjects.remove(extra_mapel)
                            if extra_mapel in st.session_state.extra_nilai:
                                del st.session_state.extra_nilai[extra_mapel]
                            if extra_mapel in st.session_state.extra_deskripsi:
                                del st.session_state.extra_deskripsi[extra_mapel]
                            st.rerun()
    
    # Gabungkan mata pelajaran tambahan ke daftar utama (untuk PDF & generate)
    final_subjects = subjects.copy()
    for extra_mapel in st.session_state.extra_subjects:
        if extra_mapel not in final_subjects:  # hindari duplikat
            final_subjects.append(extra_mapel)
            nilai_list.append(st.session_state.extra_nilai.get(extra_mapel, 75))
            deskripsi_list.append(st.session_state.extra_deskripsi.get(extra_mapel, ""))
    
    # === CAPAIAN PEMBELAJARAN (CP) & TUJUAN PEMBELAJARAN (TP) - BARU ===
    st.markdown('<div class="section-header">🎯 CAPAIAN PEMBELAJARAN (CP) & TUJUAN PEMBELAJARAN (TP)</div>', unsafe_allow_html=True)
    
    st.caption("Opsional: Ringkasan CP & TP yang menjadi fokus semester ini (bisa diambil dari ATP / Modul Ajar). Akan muncul sebagai bagian terpisah di PDF rapor untuk memperkuat kesesuaian dengan Kurikulum Merdeka.")
    
    cp_tp_ringkasan = st.text_area(
        "Ringkasan Capaian Pembelajaran & Tujuan Pembelajaran yang Dicapai Semester Ini",
        value=f"Pada semester ini peserta didik telah mencapai Capaian Pembelajaran Fase {fase} pada mata pelajaran yang diajarkan. Tujuan Pembelajaran difokuskan pada penguatan literasi, numerasi, karakter, dan Profil Pelajar Pancasila melalui pembelajaran yang bermakna dan berbasis proyek.",
        height=75,
        key="cp_tp_ringkasan",
        help="Anda dapat menyesuaikan dengan CP & TP spesifik dari dokumen ATP sekolah Anda."
    )
    
    # === DATA TAMBAHAN ===
    st.markdown('<div class="section-header">📋 DATA TAMBAHAN</div>', unsafe_allow_html=True)
    
    col1, col2, col3 = st.columns(3)
    with col1:
        sakit = st.number_input("Sakit (hari)", min_value=0, value=0, key="sakit")
    with col2:
        izin = st.number_input("Izin (hari)", min_value=0, value=1, key="izin")
    with col3:
        tanpa_keterangan = st.number_input("Tanpa Keterangan (hari)", min_value=0, value=0, key="tanpa")
    
    catatan_wali = st.text_area(
        "Catatan Wali Kelas", 
        value="Ananda menunjukkan sikap yang baik dan aktif dalam pembelajaran. Pertahankan semangat dan tingkatkan kedisiplinan serta tanggung jawab dalam mengerjakan tugas.",
        height=60,
        key="catatan_wali"
    )
    
    # Tanda tangan
    st.markdown('<div class="section-header">✍️ TANDA TANGAN</div>', unsafe_allow_html=True)
    
    col1, col2 = st.columns(2)
    with col1:
        nama_wali_kelas = st.text_input("Nama Wali Kelas", value="Siti Aminah, S.Pd.", key="wali_kelas")
        nip_wali = st.text_input("NIP Wali Kelas", value="198505152010012003", key="nip_wali")
    with col2:
        nama_kepala = st.text_input("Nama Kepala Sekolah", value="Drs. Budi Santoso, M.Pd.", key="kepala")
        nip_kepala = st.text_input("NIP Kepala Sekolah", value="196812101990011001", key="nip_kepala")
    
    tempat_tanggal = st.text_input("Tempat, Tanggal (contoh: Jakarta, 18 Juni 2026)", 
                                   value=f"{kota}, 18 Juni 2026", key="tempat_tanggal")
    
    # === GENERATE BUTTON + PILIHAN KERTAS (BARU v1.3) ===
    st.markdown("---")
    
    # Pilihan ukuran kertas sebelum generate (sangat berguna saat ada banyak mapel tambahan)
    col_paper, col_spacer = st.columns([0.55, 0.45])
    with col_paper:
        paper_choice = st.radio(
            "📄 Pilih Ukuran Kertas untuk Rapor PDF",
            options=[
                "A4 (210 × 297 mm) — Standar, paling umum digunakan",
                "F4 / Folio (210 × 330 mm) — Lebih tinggi, ideal jika banyak mata pelajaran tambahan"
            ],
            index=0,
            horizontal=False,
            help="F4 memberikan ruang vertikal ekstra ~3.3 cm sehingga tabel nilai dan tanda tangan lebih lega dan tidak cramped ketika Anda menambahkan beberapa mata pelajaran ekstra."
        )
    
    if st.button("📥 GENERATE & DOWNLOAD RAPOR PDF", type="primary", use_container_width=True):
        if not nama_siswa or not nama_sekolah:
            st.error("Nama siswa dan nama sekolah wajib diisi!")
        else:
            with st.spinner("Membuat dokumen PDF rapor..."):
                # Tentukan pagesize berdasarkan pilihan user
                if "F4" in paper_choice:
                    from reportlab.lib.units import mm
                    selected_pagesize = (210 * mm, 330 * mm)  # F4 / Legal Indonesia
                    paper_label = "F4"
                else:
                    selected_pagesize = A4
                    paper_label = "A4"
                
                data = {
                    'nama_sekolah': nama_sekolah,
                    'npsn': npsn,
                    'alamat_sekolah': alamat_sekolah,
                    'kota': kota,
                    'nama_siswa': nama_siswa,
                    'nisn': nisn,
                    'kelas': kelas,
                    'fase': fase,
                    'semester': semester,
                    'tahun_ajaran': tahun_ajaran,
                    'agama': agama,
                    'subjects': final_subjects,
                    'nilai_list': nilai_list,
                    'deskripsi_list': deskripsi_list,
                    'sakit': sakit,
                    'izin': izin,
                    'tanpa_keterangan': tanpa_keterangan,
                    'catatan_wali': catatan_wali,
                    'nama_wali_kelas': nama_wali_kelas,
                    'nip_wali': nip_wali,
                    'nama_kepala': nama_kepala,
                    'nip_kepala': nip_kepala,
                    'tempat_tanggal': tempat_tanggal,
                    'cp_tp_ringkasan': cp_tp_ringkasan,
                    # Logo terpisah untuk 3 posisi (v1.5)
                    'logo_kiri_bytes': logo_kiri.getvalue() if logo_kiri else None,
                    'logo_kanan_bytes': logo_kanan.getvalue() if logo_kanan else None,
                    'logo_wm_bytes': logo_wm.getvalue() if logo_wm else None,
                    'show_watermark': show_watermark,
                    # Warna tema & background
                    'primary_color': st.session_state.get('primary_color', '#1a5276'),
                    'table_header_bg': st.session_state.get('table_header_bg', '#d5dbdb'),
                    'row_alt_color': st.session_state.get('row_alt_color', '#f8f9f9'),
                    'page_bg_color': st.session_state.get('page_bg_color', '#FFFFFF'),
                    'kkm': kkm,
                }
                
                pdf_bytes = create_rapor_pdf(data, pagesize=selected_pagesize)
                
                filename = f"Rapor_{nama_siswa.replace(' ', '_')}_Kelas{kelas}_{semester.replace(' ', '')}_{tahun_ajaran.replace('/', '-')}_{paper_label}.pdf"
                
                st.success(f"✅ Rapor berhasil dibuat dalam ukuran **{paper_label}**! Silakan unduh di bawah ini.")
                st.download_button(
                    label=f"⬇️ UNDUH RAPOR PDF ({paper_label})",
                    data=pdf_bytes,
                    file_name=filename,
                    mime="application/pdf",
                    use_container_width=True
                )
                
                st.balloons()

# ==================== TAB 2: BATCH EXCEL ====================
with tab2:
    st.header("📦 Batch Processing - Banyak Siswa Sekaligus")
    st.markdown("""
    Fitur ini memungkinkan Anda membuat **puluhan atau ratusan rapor PDF** hanya dengan mengisi satu file Excel.
    Sangat cocok untuk sekolah yang memiliki banyak siswa per kelas.
    """)
    
    st.info("💡 **Alur kerja**: Download template Excel → Isi data & nilai siswa → Upload kembali → Generate ZIP berisi semua PDF")
    
    # Input info sekolah bersama untuk batch
    with st.expander("🏫 Isi Identitas Sekolah (berlaku untuk semua siswa dalam batch ini)", expanded=False):
        col1, col2 = st.columns(2)
        with col1:
            batch_nama_sekolah = st.text_input("Nama Sekolah Dasar", value="SD Negeri 01 Contoh", key="batch_nama_sekolah")
            batch_npsn = st.text_input("NPSN", value="12345678", key="batch_npsn")
        with col2:
            batch_alamat = st.text_input("Alamat Sekolah", value="Jl. Pendidikan No. 1", key="batch_alamat")
            batch_kota = st.text_input("Kota/Kabupaten", value="Kota Contoh", key="batch_kota")
        
        # ==================== LOGO BATCH (3 TEMPAT) - v1.5 ====================
        st.markdown("**🖼️ Upload Logo untuk Semua Rapor dalam Batch**")
        st.caption("Logo yang sama akan digunakan di semua PDF yang dihasilkan dari Excel.")
        
        col_bl1, col_bl2, col_bl3 = st.columns(3)
        
        with col_bl1:
            batch_logo_kiri = st.file_uploader(
                "📍 Logo Kiri Atas", type=["png", "jpg", "jpeg"],
                key="batch_logo_kiri",
                help="Logo pojok kiri atas header"
            )
        with col_bl2:
            batch_logo_kanan = st.file_uploader(
                "📍 Logo Kanan Atas", type=["png", "jpg", "jpeg"],
                key="batch_logo_kanan",
                help="Logo pojok kanan atas header"
            )
        with col_bl3:
            batch_logo_wm = st.file_uploader(
                "💧 Logo Watermark Tengah", type=["png", "jpg", "jpeg"],
                key="batch_logo_wm",
                help="Logo besar untuk watermark tengah"
            )
        
        # Watermark checkbox untuk batch
        batch_show_watermark = st.checkbox(
            "💧 Tampilkan Logo Watermark Tengah di TENGAH halaman (untuk semua siswa)",
            value=True if batch_logo_wm else False,
            key="batch_show_watermark",
            disabled=batch_logo_wm is None,
            help="Logo watermark akan muncul samar di tengah setiap PDF dalam ZIP."
        )
        
        # ==================== TEMA WARNA UNTUK BATCH (sama untuk semua PDF dalam ZIP) ====================
        st.markdown("---")
        st.subheader("🎨 Tema Warna untuk Semua Rapor dalam Batch")
        batch_theme_options = list(COLOR_THEMES.keys()) + ["Custom (Pilih Warna Manual)"]
        batch_theme_choice = st.selectbox(
            "Pilih Tema Warna (berlaku untuk semua siswa)",
            options=batch_theme_options,
            index=0,
            key="batch_theme_choice"
        )
        
        if batch_theme_choice == "Custom (Pilih Warna Manual)":
            col_bc1, col_bc2 = st.columns(2)
            with col_bc1:
                batch_primary = st.color_picker("Warna Utama Header/Border", "#1a5276", key="batch_custom_primary")
                batch_table_header = st.color_picker("Warna Header Tabel", "#d5dbdb", key="batch_custom_table")
            with col_bc2:
                batch_row_alt = st.color_picker("Warna Baris Selang-seling", "#f8f9f9", key="batch_custom_row")
                batch_page_bg = st.color_picker("Background Halaman", "#FFFFFF", key="batch_custom_page")
        else:
            btheme = COLOR_THEMES[batch_theme_choice]
            batch_primary = btheme["primary"]
            batch_table_header = btheme["table_header_bg"]
            batch_row_alt = btheme["row_alt"]
            batch_page_bg = "#FFFFFF"
            st.caption(f"Tema: {batch_theme_choice} — {btheme['description']}")
        
        # Simpan ke session untuk digunakan saat generate batch
        st.session_state['batch_primary_color'] = batch_primary
        st.session_state['batch_table_header_bg'] = batch_table_header
        st.session_state['batch_row_alt_color'] = batch_row_alt
        st.session_state['batch_page_bg_color'] = batch_page_bg
    
    # Tombol download template
    if st.button("📥 DOWNLOAD TEMPLATE EXCEL SIAP PAKAI", type="primary", use_container_width=True):
        template_bytes = create_excel_template()
        st.download_button(
            label="⬇️ Unduh Template_Rapor_Batch.xlsx",
            data=template_bytes,
            file_name="Template_Rapor_Batch_SD_Kelas1-6.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True
        )
        st.success("Template berhasil dibuat! Buka file Excel tersebut dan isi data siswa.")
    
    st.markdown("---")
    
    # Upload Excel
    uploaded_file = st.file_uploader(
        "Upload File Excel yang sudah diisi",
        type=["xlsx", "xls"],
        help="File harus mengikuti format template di atas"
    )
    
    if uploaded_file is not None:
        try:
            df = pd.read_excel(uploaded_file, sheet_name="Data_Siswa")
            df = df.dropna(how='all')  # hapus baris kosong
            
            st.success(f"✅ File berhasil dibaca! Ditemukan **{len(df)} baris data**.")
            
            # Preview
            with st.expander("👀 Lihat Preview Data (10 baris pertama)", expanded=True):
                st.dataframe(df.head(10), use_container_width=True)
            
            # Tombol generate
            if st.button("🚀 GENERATE SEMUA PDF & DOWNLOAD ZIP", type="primary", use_container_width=True):
                with st.spinner(f"Memproses {len(df)} siswa... Mohon tunggu beberapa saat"):
                    zip_bytes, success_count, error_list = process_batch_excel(
                        df,
                        nama_sekolah=batch_nama_sekolah,
                        npsn=batch_npsn,
                        alamat_sekolah=batch_alamat,
                        kota=batch_kota,
                        logo_kiri_bytes=batch_logo_kiri.getvalue() if batch_logo_kiri else None,
                        logo_kanan_bytes=batch_logo_kanan.getvalue() if batch_logo_kanan else None,
                        logo_wm_bytes=batch_logo_wm.getvalue() if batch_logo_wm else None,
                        muatan_lokal_name=batch_muatan_lokal,
                        show_watermark=batch_show_watermark,
                        primary_color=st.session_state.get('batch_primary_color', '#1a5276'),
                        table_header_bg=st.session_state.get('batch_table_header_bg', '#d5dbdb'),
                        row_alt_color=st.session_state.get('batch_row_alt_color', '#f8f9f9'),
                        page_bg_color=st.session_state.get('batch_page_bg_color', '#FFFFFF')
                    )
                
                st.success(f"🎉 Selesai! **{success_count}** rapor berhasil dibuat.")
                
                if error_list:
                    st.warning(f"Ada {len(error_list)} error. Lihat detail di file RINGKASAN_GENERATE.txt di dalam ZIP.")
                
                # Download ZIP
                zip_filename = f"Rapor_Batch_{datetime.now().strftime('%Y%m%d_%H%M')}.zip"
                st.download_button(
                    label=f"⬇️ UNDUH {success_count} FILE PDF (ZIP)",
                    data=zip_bytes,
                    file_name=zip_filename,
                    mime="application/zip",
                    use_container_width=True
                )
                
                # Tampilkan ringkasan error jika ada
                if error_list:
                    with st.expander("Lihat Daftar Error"):
                        for err in error_list:
                            st.error(err)
        
        except Exception as e:
            st.error(f"Terjadi kesalahan saat membaca file Excel: {str(e)}")
            st.info("Pastikan file yang diupload adalah template yang benar dan sheet bernama 'Data_Siswa'.")

# ==================== TAB 3: PANDUAN ====================
with tab3:
    st.header("📖 Panduan Penggunaan & Catatan Penting")
    
    st.markdown("""
    ### Fitur Utama Aplikasi
    - **Input mudah** untuk semua mata pelajaran SD (Kelas 1-6)
    - **Deskripsi otomatis** yang positif, memotivasi, dan sesuai bahasa Kurikulum Merdeka
    - **Predikat otomatis** (A/B/C/D) berdasarkan nilai
    - **PDF profesional** siap cetak dengan layout rapi
    - Mendukung **Fase A (Kelas 1-2)**, **Fase B (Kelas 3-4)**, **Fase C (Kelas 5-6)**
    
    ### Mata Pelajaran yang Didukung
    | Kelas | Mata Pelajaran Utama |
    |-------|----------------------|
    | 1-2   | Agama & Budi Pekerti, Pendidikan Pancasila, Bahasa Indonesia, Matematika, PJOK, Seni Budaya, Muatan Lokal |
    | 3-4   | + IPAS |
    | 5-6   | + Bahasa Inggris |
    
    ### Tips Pengisian yang Akurat
    1. Nilai Akhir adalah nilai akhir/holistik (bisa rata-rata Pengetahuan + Keterampilan atau penilaian menyeluruh)
    2. Edit deskripsi sesuai observasi nyata siswa (deskripsi otomatis hanya panduan)
    3. Gunakan bahasa positif dan memotivasi seperti contoh resmi Kemendikdasmen
    4. Untuk siswa dengan nilai rendah, deskripsi tetap memberikan harapan dan arahan perbaikan
    
    ### Format PDF
    - Ukuran A4 Portrait
    - Layout mengikuti contoh resmi Panduan Pembelajaran dan Asesmen
    - Siap untuk dicetak atau diarsipkan secara digital
    - Tanda tangan digital-ready (bisa ditambah tanda tangan scan jika diedit lebih lanjut)
    
    ### Batasan & Saran
    - Aplikasi ini untuk **satu siswa per generate** (ulangi untuk siswa lain)
    - Untuk sekolah besar, Anda bisa membuat script batch tambahan atau gunakan Excel sebagai sumber data
    - Logo sekolah belum disertakan di PDF versi ini (bisa ditambahkan di pengembangan selanjutnya)
    - Simpan file PDF dengan nama yang jelas untuk arsip
    
    ### Legal & Akurasi
    Format ini mengacu pada:
    - Panduan Pembelajaran dan Asesmen Pendidikan Dasar (Kemendikdasmen)
    - Contoh Format Laporan Hasil Belajar (Rapor) Jenjang SD Fase A, B, dan C
    - Struktur Kurikulum SD Permendikdasmen terbaru
    
    Anda tetap bertanggung jawab atas keakuratan data dan deskripsi yang diinput.
    """)
    
    st.success("Aplikasi ini dibuat untuk memudahkan guru SD dalam menyusun rapor yang akurat, rapi, dan sesuai standar Kurikulum Merdeka.")

# Footer
st.markdown("---")
st.caption("© 2026 Aplikasi Rapor SD | Dibuat dengan ❤️ untuk kemudahan administrasi guru Indonesia")